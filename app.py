import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Wholesale Markup Analytics", layout="wide")
st.title("💰 Wholesale Markup Analytics Tool")

@st.cache_data
def load_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    return pd.read_excel(file)

st.header("Upload Files")

inv_file = st.file_uploader("Invoices")
prod_file = st.file_uploader("Products File")
front_file = st.file_uploader("Frontline")
tax_file = st.file_uploader("Taxes")
store_file = st.file_uploader("Storelist")

st.markdown("### 📊 Markup % Formula")
st.info("Markup % = (Invoice Cost - Total Cost) / Total Cost")

if inv_file and prod_file and front_file and tax_file and store_file:

    inv = load_file(inv_file)
    prod = load_file(prod_file)
    front = load_file(front_file)
    tax = load_file(tax_file)
    store = load_file(store_file)

    st.success("Files loaded")

    # ==============================
    # COLUMN MAPPING
    # ==============================
    inv_store = "store"
    inv_product = "productId"
    inv_cost = "price"

    prod_id = "ProductId"
    prod_family = "Family"
    prod_type = "Type"
    prod_case = "Products/Case"

    front_family = "Family"
    front_cost = "CasePrice"
    front_start = "Start"
    front_end = "End"

    store_store = "uniqueId"
    store_state = "stateAbbrev"

    # TAX SELECTORS
    tax_state = st.selectbox("Tax State", tax.columns)
    tax_type_col = st.selectbox("Tax Product Type", tax.columns)
    tax_percentage_col = st.selectbox("Percentage", tax.columns)
    tax_value = st.selectbox("Tax", tax.columns)
    uom_tax_col = st.selectbox("Products/Case * Tax", tax.columns)

    if st.button("🚀 Run Analysis"):

        progress = st.progress(0)

        # ==============================
        # HELPERS
        # ==============================
        def clean_numeric(series):
            return pd.to_numeric(series.astype(str).str.strip(), errors="coerce")

        def clean_text(series):
            return series.astype(str).str.strip().str.upper()

        def clean_id(x):
            return str(x).strip().lstrip("0")

        # ==============================
        # CLEAN KEYS
        # ==============================
        inv["ProductID"] = inv[inv_product].apply(clean_id)
        prod["ProductID"] = prod[prod_id].apply(clean_id)

        inv["Store"] = inv[inv_store].astype(str).str.strip()
        store["Store"] = store[store_store].astype(str).str.strip()

        store["State"] = store[store_state].astype(str).str.strip()
        tax["State"] = tax[tax_state].astype(str).str.strip()

        prod["Family"] = clean_text(prod[prod_family])
        front["Family"] = clean_text(front[front_family])

        progress.progress(10)

        # ==============================
        # PRODUCT DEDUP
        # ==============================
        prod = prod.drop_duplicates(subset=["ProductID"])

        # ==============================
        # MERGE PRODUCT
        # ==============================
        merged = inv.merge(
            prod[["ProductID", "Family", "Type", prod_case]],
            on="ProductID",
            how="left"
        )

        progress.progress(25)

        # ==============================
        # NORMALIZE TYPE
        # ==============================
        merged["Type"] = clean_text(merged["Type"])
        tax["ProductType"] = clean_text(tax[tax_type_col])

        # ==============================
        # ACTIVE FRONTLINE
        # ==============================
        today = pd.Timestamp.today().normalize()

        front[front_start] = pd.to_datetime(front[front_start], errors="coerce")
        front[front_end] = pd.to_datetime(front[front_end], errors="coerce")
        front[front_end] = front[front_end].fillna(pd.Timestamp.max)

        active_front = front[
            (front[front_start] <= today) &
            (front[front_end] >= today)
        ].copy()

        active_front = (
            active_front
            .sort_values(front_start, ascending=False)
            .drop_duplicates(subset=["Family"])
        )

        progress.progress(45)

        # ==============================
        # MERGE FRONTLINE
        # ==============================
        merged = merged.merge(
            active_front[["Family", front_cost]],
            on="Family",
            how="left"
        )

        progress.progress(60)

        # ==============================
        # STORE → STATE
        # ==============================
        merged = merged.merge(
            store[["Store", "State"]],
            on="Store",
            how="left"
        )

        progress.progress(75)

        # ==============================
        # TAX MERGE (STATE + TYPE)
        # ==============================
        merged = merged.merge(
            tax,
            left_on=["State", "Type"],
            right_on=["State", "ProductType"],
            how="left"
        )

        progress.progress(85)

        # ==============================
        # CLEAN NUMERIC
        # ==============================
        merged["Percentage"] = clean_numeric(merged[tax_percentage_col])
        merged["TaxValue"] = clean_numeric(merged[tax_value])
        merged["Products/Case * Tax"] = clean_numeric(merged[uom_tax_col])
        merged["Frontline"] = clean_numeric(merged[front_cost])
        merged["Products/Case"] = clean_numeric(merged["Products/Case"])

        # ==============================
        # TAX ENGINE
        # ==============================
        merged["Tax"] = 0.0
        merged["Tax Rule Applied"] = "None"

        # 1️⃣ Percentage
        mask_pct = merged["Percentage"].notna()
        merged.loc[mask_pct, "Tax"] = merged["Frontline"] * merged["Percentage"]
        merged.loc[mask_pct, "Tax Rule Applied"] = "Percentage"

        # 2️⃣ Tax column
        mask_tax = merged["TaxValue"].notna() & (~mask_pct)
        merged.loc[mask_tax, "Tax"] = merged["TaxValue"]
        merged.loc[mask_tax, "Tax Rule Applied"] = "Tax Value"

        # 3️⃣ Case rate
        mask_case = (
            merged["Products/Case * Tax"].notna()
        ) & (~mask_pct) & (~mask_tax)

        merged.loc[mask_case, "Tax"] = (
            merged["Products/Case"] * merged["Products/Case * Tax"]
        )
        merged.loc[mask_case, "Tax Rule Applied"] = "Case * Rate"

        merged["Tax"] = merged["Tax"].fillna(0)

        # ==============================
        # CALCULATIONS
        # ==============================
        merged["Invoice Cost"] = clean_numeric(merged[inv_cost])

        merged["Total Cost"] = merged["Frontline"] + merged["Tax"]
        merged["Markup"] = merged["Invoice Cost"] - merged["Total Cost"]
        merged["Markup %"] = merged["Markup"] / merged["Total Cost"]

        merged["Markup %"] = merged["Markup %"].replace([float("inf"), -float("inf")], 0)

        progress.progress(90)

        # ==============================
        # FREQUENCY
        # ==============================
        freq = (
            merged
            .groupby(["State", "Family", "Type", "Invoice Cost"])
            .size()
            .reset_index(name="Frequency")
        )

        freq["Top"] = (
            freq.groupby(["State", "Family", "Type"])["Frequency"]
            .transform("max") == freq["Frequency"]
        )

        merged = merged.merge(freq, on=["State", "Family", "Type", "Invoice Cost"], how="left")

        # ==============================
        # DEDUP
        # ==============================
        merged = merged.sort_values("Tax", ascending=False)

        merged = merged.drop_duplicates(
            subset=["State", "Family", "Type", "Invoice Cost"],
            keep="first"
        )

        # ==============================
        # FINAL OUTPUT
        # ==============================
        final = merged[[
            "State","Family","Type","Invoice Cost","Frontline","Tax",
            "Total Cost","Markup","Markup %","Frequency","Top","Tax Rule Applied"
        ]]

        # ==============================
        # EXPORT WITH HIGHLIGHT
        # ==============================
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            final.to_excel(writer, sheet_name="Analysis", index=False)
            merged.to_excel(writer, sheet_name="Full Output", index=False)

        output.seek(0)

        wb = load_workbook(output)
        ws = wb["Analysis"]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        top_col_index = list(final.columns).index("Top") + 1

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=top_col_index).value:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = green_fill

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button(
            "📥 Download Analysis",
            data=final_output,
            file_name=f"markup_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
